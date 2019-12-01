from openpyxl import Workbook,load_workbook

#wb=Workbook("testxl.xlsx") create new workbook

wb=load_workbook("testxl.xlsx") #load existing workbook

wb.create_sheet(title="test1",index=0) # create worksheet

print wb.sheetnames #names of present sheet


wb["Sheet"].title="test0" #change name of sheet

 wb.remove_sheet(wb["test0"]) #delete sheet



sheet=wb["test1"] # or wb.active  will load sheet

row=[["Name","Place","Animal","Thing"]] 

for x in row:
	sheet.append(x)  # will append value in row i.e last row *note: values should be in tuple/list

name=["ameya","bob","cathy","dan"]

c=sheet.max_row #shows how many rows are there

for x in name:
	sheet.cell(row=c,column=1,value=x)	#insert values in particular column
	c+=1 									

b=sheet.max_column #shows how many columns are there

title=["Name","Place","Animal","Thing"]

for x in title:
	sheet.cell(row=c,column=b,value=x) #insert values in particular row
	b+=1


wb.save("testxl.xlsx") #save it to confirm











wb.save("testxl.xlsx")
